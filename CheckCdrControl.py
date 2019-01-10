import openpyxl, sys
import time
import datetime
import os
from tkinter import messagebox
import time

class CheckCdrControl:

    siteWithError=[]

    def __init__(self, paths):
        # super(CheckCdrControl, self).__init__()
        self.paths = paths
        self.errInCDR=False
        self.listLineerrInCDR=[]
        self.siteWithError=[]
        self.exc_type=""
        self.exc_obj=""
        self.exc_tb=""
        self.fileWithGenericError=""

        localErr=False
        for element in self.paths:
            self.errInCDR=False
            if element!= None:
                if "GSM" in element or"gsm"in element or "2g"in element or "2G"in element:
                    # try:
                    self.gsm(element)
                    # except Exception as e:
                    #     self.errInCDR=None
                    #     self.exc_type, self.exc_obj, self.exc_tb = sys.exc_info()
                    #     self.fileWithGenericError=element.split("/")[-1]
                    #     break
                    if self.errInCDR==True:
                        self.siteWithError.append(element.split("/")[-1])
                        self.createLog(element)
                        localErr=True
                        continue
                    if self.errInCDR==None:
                        break
                elif "WCDMA"in element or "wcdma" in element or "3g" in element or "3G" in element or "umts" in element or "UMTS" in element:
                    # try:
                    self.wcdma(element)
                    # except Exception as e:
                    #     self.errInCDR=None
                    #     self.exc_type, self.exc_obj, self.exc_tb = sys.exc_info()
                    #     self.fileWithGenericError=element.split("/")[-1]
                    #     break
                    if self.errInCDR==True:
                        self.siteWithError.append(element.split("/")[-1])
                        self.createLog(element)
                        localErr=True
                        continue
                    elif self.errInCDR==None:
                        break
                elif "LTE" in element or "lte" in element or "4g"in element or "4G"in element:
                    # try:
                    self.lte(element)
                    # except Exception as e:
                    #     self.errInCDR=None
                    #     self.exc_type, self.exc_obj, self.exc_tb = sys.exc_info()
                    #     self.fileWithGenericError=element.split("/")[-1]
                    #     break
                    if self.errInCDR==True:
                        self.siteWithError.append(element.split("/")[-1])
                        self.createLog(element)
                        localErr=True
                        continue
                    elif self.errInCDR==None:
                        break
                else:
                    messagebox.showerror(message=f"The file {element} must descript in the name, the tipology of CDR(2g, 3g, 4g)")
                pass


        if localErr==True:
            joinSitewithError=",".join(self.siteWithError)
            error=messagebox.askyesno(title="Warning",message=f"There are errors in CDR {joinSitewithError}. Do you want open the log folder?")
            if error:
                os.startfile(".\\Log Error In Cdr")
        elif  self.errInCDR==False:
            messagebox.showinfo(message="There aren't errors")
        # elif self.errInCDR=="error ":
        #     messagebox.showerror(message=f"Generic error in Tool with {self.fileWithGenericError}.\nError type: {self.exc_type}\nError line: {self.exc_tb.tb_lineno}")



    def createLog(self, path):
        pathSplit=path.split("/")
        nameFile=pathSplit[-1]
        if not os.path.exists("./Log Error In Cdr"):
            os.makedirs("./Log Error In Cdr")
        file=open("./Log Error In Cdr/Log_"+nameFile+"_"+str(datetime.datetime.fromtimestamp(time.time()).strftime('%m_%d %H_%M'))+".txt","w")
        file.writelines(self.listLineerrInCDR)
        self.listLineerrInCDR=[]
        file.close()


        pass

    def gsm(self, path):
        try:
            doc= openpyxl.load_workbook(path, data_only=True)
            nTotCol=countCol(doc,"BTS","C",2)
            elementColE=elemInCol(doc,"BTS","E",nTotCol,2)

            listRecurrence=[]

            cellCreated=[]
            frqBCCH=[]
            for element in range(2,nTotCol+2):
                cellCreated.append(doc["BTS"]["C"+str(element)].value)
                frqBCCH.append(doc["BTS"]["K"+str(element)].value)

            # in questo for controllo se ci sono valori ripetuti in H
            for element in elementColE:

                recurrence=0
                for elementsToCheck in elementColE:
                    if element==elementsToCheck:
                        recurrence+=1
                    pass
                pass

                if recurrence>1:
                    self.errInCDR=True
                    if (element in listRecurrence)==False:
                        listRecurrence.append(element)
                        self.listLineerrInCDR.append("The \""+element+"\" is present more than once in column \"TG\" in sheet \"BTS\"\n\n")
                pass

            # in questo for faccio diversi controlli su alcuni campi
            for element in range(2,nTotCol+2):
                cellName=doc["BTS"]["C"+str(element)].value

                if cellName[-2]=="G":
                    if doc["BTS"]["K"+str(element)].value<100 or doc["BTS"]["K"+str(element)].value>124:
                        self.errInCDR=True
                        self.listLineerrInCDR.append("The value in \""+doc["BTS"]["K"+str(1)].value+"\"(K"+str(element)+") must be between 100 and 124\n\n")
                elif cellName[-2]=="D":
                    if doc["BTS"]["K"+str(element)].value<687 or doc["BTS"]["K"+str(element)].value>710:
                        self.errInCDR=True
                        self.listLineerrInCDR.append("The value in \""+doc["BTS"]["K"+str(1)].value+"\"(K"+str(element)+") must be between 687 and 710\n\n")


                # controllo che se il valoe di h ci devono stare f e g
                if not(str(doc["BTS"]["F"+str(element)].value) in str(doc["BTS"]["H"+str(element)].value)) or not(str(doc["BTS"]["G"+str(element)].value) in str(doc["BTS"]["H"+str(element)].value)):
                    self.errInCDR=True
                    self.listLineerrInCDR.append("The format of \""+str(doc["BTS"]["H"+str(element)].value) + "\" in column \""+str(doc["BTS"]["H"+str(1)].value)+"\"(H"+str(element)+") in sheet \"BTS\", is wrong. "+
                                               "It must have inside \""+ str(doc["BTS"]["F"+str(element)].value)+ "\""+"(F"+str(element)+") and \""+str(doc["BTS"]["G"+str(element)].value)+"\"(G"+str(element)+")\n\n")

                # Diversi controlli sul TRX in caso non fosse vuoto o diverso da 0
                if (doc["BTS"]["X"+str(element)].value!=None) and (doc["BTS"]["X"+str(element)].value!=0):

                    nTrx=doc["BTS"]["X"+str(element)].value
                    startLett="F"
                    startLett1="C"
                    cellName=doc["BTS"]["C"+str(element)].value
                    # se il valore in HSN (CHGR-1) (AB) è vuoto restituisco errore
                    if (doc["BTS"]["AB"+str(element)].value==None):
                        self.errInCDR=True
                        self.listLineerrInCDR.append("The value in \""+doc["BTS"]["AB"+str(1)].value+"\"(AB"+str(element)+") in sheet \"BTS\" must not be empty"+"\n\n")
                    else:
                        # print(doc["BTS"]["AB"+str(element)].value)
                        if int(doc["BTS"]["AB"+str(element)].value)<0 or int(doc["BTS"]["AB"+str(element)].value)>63:
                            self.errInCDR=True
                            self.listLineerrInCDR.append("The value in \""+doc["BTS"]["AB"+str(1)].value+"\"(AB"+str(element)+") must be between 0 and 63.\n\n")

                    # se la penultima lettera di cellname(C) è G allora anche rSite(D) deve finire in d, stessa cosa per Cellname(c)con G
                    if cellName[-2]=="G":
                        if doc["BTS"]["D"+str(element)].value[-1]!="G":
                            self.errInCDR=True
                            self.listLineerrInCDR.append("The value in \""+doc["BTS"]["D"+str(1)].value+"\"(D"+str(element)+") in sheet \"BTS\" must end with \"G\""+"\n\n")
                    elif cellName[-2]=="D":
                        if doc["BTS"]["D"+str(element)].value[-1]!="D":
                            self.errInCDR=True
                            self.listLineerrInCDR.append("The value in \""+doc["BTS"]["D"+str(1)].value+"\"(D"+str(element)+") in sheet \"BTS\" must end with \"D\""+"\n\n")

                    # i maio devono essere compilati almeno per un numero pari ai trx e il loro valore oscilla tra 0 e il numero di colonne tchFreq compilate
                    for e in range(0, nTrx):
                        if doc["BTS"]["A"+startLett1+str(element)].value==None:
                            self.errInCDR=True
                            self.listLineerrInCDR.append("The cell in \""+doc["BTS"]["A"+startLett1+str(1)].value+"\"(A"+startLett1+str(element)+") in sheet \"BTS\" "+"can not be empty.\n\n")
                        elif doc["BTS"]["A"+startLett1+str(element)].value!=None:
                            valMax=countRowFill(doc, element, "AF")
                            if int(doc["BTS"]["A"+startLett1+str(element)].value)>int(valMax) or int(doc["BTS"]["A"+startLett1+str(element)].value)<0:
                                self.errInCDR=True
                                self.listLineerrInCDR.append("The value in \""+doc["BTS"]["A"+startLett1+str(1)].value+"\"(A"+startLett1+str(element)+") must be between 0 and "+str(valMax)+"\n\n")
                        startLett1=chr(ord(startLett1)+1)

                    # i tchFreq devono essere compilati per un numero almeno pari a trx + 1 e
                    # il loro vvalore oscilla tra 687 e 710 in caso cellname sia D altrimenti tra 100 e 124 in caso cell name sia d
                    # il loro valore non può essere presente nella lista frqBCCH
                    for e in range(0,nTrx+1):

                        if doc["BTS"]["A"+startLett+str(element)].value==None:
                            self.errInCDR=True
                            self.listLineerrInCDR.append("The cell in "+doc["BTS"]["A"+startLett+str(1)].value+"(A"+startLett+str(element)+") in sheet \"BTS\" "+"can not be empty.\n\n")
                        else:
                            if cellName[-2]=="G":
                                if int(doc["BTS"]["A"+startLett+str(element)].value)<100 or int(doc["BTS"]["A"+startLett+str(element)].value)>124:
                                    self.errInCDR=True
                                    self.listLineerrInCDR.append("The value in \""+doc["BTS"]["A"+startLett+str(1)].value+"\"(A"+startLett+str(element)+") must be between 100 and 124\n\n")

                            elif cellName[-2]=="D":
                                if int(doc["BTS"]["A"+startLett+str(element)].value)<687 or int(doc["BTS"]["A"+startLett+str(element)].value)>710:
                                    self.errInCDR=True
                                    self.listLineerrInCDR.append("The value in \""+doc["BTS"]["A"+startLett+str(1)].value+"\"(A"+startLett+str(element)+") must be between 687 and 710\n\n")

                        startLett=chr(ord(startLett)+1)
                # se TRX è vuota controllo che da af a ak della stessa colonna siano vuote
                elif (doc["BTS"]["X"+str(element)].value==None) or (doc["BTS"]["X"+str(element)].value==0):

                    errTrxEmpty=False

                    if doc["BTS"]["AF"+str(element)].value!=None:
                        errTrxEmpty=True
                    if doc["BTS"]["AG"+str(element)].value!=None:
                        errTrxEmpty=True
                    if doc["BTS"]["AH"+str(element)].value!=None:
                        errTrxEmpty=True
                    if doc["BTS"]["AI"+str(element)].value!=None:
                        errTrxEmpty=True
                    if doc["BTS"]["AJ"+str(element)].value!=None:
                        errTrxEmpty=True
                    if doc["BTS"]["AK"+str(element)].value!=None:
                        errTrxEmpty=True

                    if errTrxEmpty==True:
                        self.errInCDR=True
                        self.listLineerrInCDR.append("The \"Number of TRX of TCH\"(X"+str(element)+")is empty and some field from coloumn \"tchFreq-0 (CHGR-1)\"(AF) to \"tchFreq-5 (CHGR-1)\"(AK) are filled\n\n")

                #controllo che la frequenza k non sia in nessuna frequenza da af a ak
                if (doc["BTS"]["K"+str(element)].value!=None) and (doc["BTS"]["K"+str(element)].value!=0):
                    valInK=doc["BTS"]["K"+str(element)].value
                    startLett="F"
                    while startLett!="L":
                        for cont in range(2,nTotCol+2):
                            if valInK==doc["BTS"]["A"+startLett+str(cont)].value:
                                self.errInCDR=True
                                self.listLineerrInCDR.append("The frequence in \"bcchFreq (CHGR-0)\"(K"+str(element)+") is present also in column \""+doc["BTS"]["A"+startLett+str(1)].value+"\"(A"+startLett+")\n\n")
                        startLett=chr(ord(startLett)+1)

                # Se il campo BSPWRB p sono vuoti restituisco errore
                if doc["BTS"]["P"+str(element)].value==None:
                    self.errInCDR=True
                    self.listLineerrInCDR.append("The cell in \""+doc["BTS"]["P"+str(1)].value+"\"(P"+str(element)+") in sheet \"BTS\" "+"can not be empty.\n\n")
                else:
                    if doc["BTS"]["P"+str(element)].value<0 or doc["BTS"]["P"+str(element)].value>43:
                        cell=worksheet[f"{row}"].value
                        # nameCol=worksheet[f"{}"].value
                        self.errInCDR=True
                        self.listLineerrInCDR.append(f"The cell in \""+doc["BTS"]["P"+str(1)].value+"\"(P"+str(element)+") in sheet \"BTS\" "+"must be between 0 and 43.\n\n")

                # il valore di Q non può essere vuoto
                if doc["BTS"]["Q"+str(element)].value==None:
                    self.errInCDR=True
                    self.listLineerrInCDR.append("The cell in \""+doc["BTS"]["Q"+str(1)].value+"\"(Q"+str(element)+") in sheet \"BTS\" "+"can not be empty.\n\n")
                else:
                    if doc["BTS"]["Q"+str(element)].value<0 or doc["BTS"]["Q"+str(element)].value>43:
                        cell=worksheet[f"{row}"].value
                        # nameCol=worksheet[f"{}"].value
                        self.errInCDR=True
                        self.listLineerrInCDR.append(f"The cell in \""+doc["BTS"]["Q"+str(1)].value+"\"(Q"+str(element)+") in sheet \"BTS\" "+"must be between 0 and 43.\n\n")

                # Se il campo BSPWR r sono vuoti restituisco errore
                if doc["BTS"]["R"+str(element)].value==None:
                    self.errInCDR=True
                    self.listLineerrInCDR.append("The cell in \""+doc["BTS"]["R"+str(1)].value+"\"(R"+str(element)+") in sheet \"BTS\" "+"can not be empty.\n\n")
                # Se il campo BSPWR r sono vuoti restituisco errore
                if doc["BTS"]["S"+str(element)].value==None:
                    self.errInCDR=True
                    self.listLineerrInCDR.append("The cell in \""+doc["BTS"]["S"+str(1)].value+"\"(S"+str(element)+") in sheet \"BTS\" "+"can not be empty.\n\n")

            colWithDiffInSameG2U=equalValuesInSameCol(doc,"ADJ G2U",["A", "B"],2)
            if colWithDiffInSameG2U:
                for element in colWithDiffInSameG2U:
                    nameCol= doc["ADJ G2U"][element+str(1)].value
                    self.errInCDR=True
                    self.listLineerrInCDR.append("There are values different in column \""+ str(nameCol) + "\" in sheet \"ADJ G2U\"\n\n")

            colWithDiffInSameG2G=equalValuesInSameCol(doc,"ADJ G2G",["A", "B"],2)
            if colWithDiffInSameG2G:
                for element in colWithDiffInSameG2G:
                    nameCol= doc["ADJ G2G"][element+str(1)].value
                    self.errInCDR=True
                    self.listLineerrInCDR.append("There are values different in column \""+ str(nameCol) + "\" in sheet \"ADJ G2G\"\n\n")

            colWithDiffInSameBTS=equalValuesInSameCol(doc,"BTS",["B"],2)
            if colWithDiffInSameBTS:
                for element in colWithDiffInSameBTS:
                    nameCol= doc["BTS"][element+str(1)].value
                    self.errInCDR=True
                    self.listLineerrInCDR.append("There are values different in column \""+ str(nameCol) + "\" in sheet \"BTS\"\n\n")

            colWithDiffInOtherG2U=equalValuesInOtherCol(doc, "ADJ G2U", "B", 2, "BTS", "B", 2)
            if colWithDiffInOtherG2U:
                nameColBTS= doc["BTS"]["B1"].value
                nameColG2U= doc["ADJ G2U"]["B1"].value
                self.errInCDR=True
                self.listLineerrInCDR.append("There are values different between column \""+ str(nameColBTS) + "\" in sheet \"BTS\" and column \""+ str(nameColG2U) + "\" in sheet \"ADJ G2U\"\n\n")

            colWithDiffInOtherG2G=equalValuesInOtherCol(doc, "ADJ G2G", "B", 2, "BTS", "B", 2)
            if colWithDiffInOtherG2G:
                nameColBTS= doc["BTS"]["B1"].value
                nameColG2G= doc["ADJ G2G"]["B1"].value
                self.errInCDR=True
                self.listLineerrInCDR.append("There are values different between column \""+ str(nameColBTS) + "\" in sheet \"BTS\" and column \""+ str(nameColG2G) + "\" in sheet \"ADJ G2G\"\n\n")


            worksheet=doc["ADJ G2G"]
            nTotRow=countCol(doc,"ADJ G2G", "B", 2 )
            for row in range(2, nTotRow+2):
                if not(worksheet[f"C{row}"].value in cellCreated):
                    cell=worksheet[f"C{row}"].value
                    nameCol=worksheet[f"C{1}"].value
                    self.errInCDR=True
                    self.listLineerrInCDR.append(f"The  cell {cell}, in {nameCol}(C{row}) in \"ADJ G2G\", is not created in column \"cell-name\"(C), in sheet \"BTS\".\n\n")
                if not(worksheet[f"E{row}"].value in cellCreated):
                    cell=worksheet[f"E{row}"].value
                    nameCol=worksheet[f"E{1}"].value
                    self.errInCDR=True
                    self.listLineerrInCDR.append(f"The  cell {cell}, in {nameCol}(E{row}) in \"ADJ G2G\", is not created in column \"cell-name\"(C), in sheet \"BTS\".\n\n")

            worksheet=doc["External 3G cells"]
            nTotRow=countCol(doc,"External 3G cells", "B", 2 )
            cell3G=[]
            for row in range(2, nTotRow+2):
                cell3G.append(worksheet[f"B{row}"].value)

            worksheet=doc["ADJ G2U"]
            nTotRow=countCol(doc,"ADJ G2U", "D", 2 )
            for row in range(2, nTotRow+2):
                if not(worksheet[f"D{row}"].value in cell3G):
                    cell=worksheet[f"D{row}"].value
                    nameCol=worksheet[f"D{1}"].value
                    self.errInCDR=True
                    self.listLineerrInCDR.append(f"The cell {cell}(D{row}) in {nameCol} in sheet \"ADJ G2U\", not exists in column \"TARGET_3G_CELL_NAME\" in sheet \"External 3G cells\".\n\n")
                    # print(f"The cell {cell}(D{row}) in {nameCol} in sheet \"ADJ G2U\", not exists in column \"TARGET_3G_CELL_NAME\" in sheet \"External 3G cells\".\n\n")


            doc.close()

        except Exception as e:
            self.errInCDR=None
            self.exc_type, self.exc_obj, self.exc_tb = sys.exc_info()
            self.fileWithGenericError=path.split("/")[-1]
            messagebox.showerror(message=f"Generic error in Tool with {self.fileWithGenericError}.\nError type: {self.exc_type}\nError line: {self.exc_tb.tb_lineno}")
        pass

    def wcdma(self, path):
        try:
            doc= openpyxl.load_workbook(path, data_only=True)
            rncDataset=doc["RNC Dataset-1"]["C3"].value
            rbsDataset=doc["RBS Dataset-1"]["C3"].value
            uniqueCode=doc["RBS Dataset-1"]["E3"].value
            nSector=nSectors(uniqueCode)

            cellCreated=[]
            lettercellCreated=[]
            worksheet=doc["RN RNC-RBS Dataset-1"]
            nTotRow=countCol(doc,"RN RNC-RBS Dataset-1", "R", 12 )
            for row in range(12, nTotRow+12):
                valColR=worksheet[f"R{row}"].value

                if not(valColR in cellCreated):
                    cellCreated.append(str(valColR))
                letterCell=valColR[-2]
                if not(letterCell in lettercellCreated):
                    lettercellCreated.append(str(letterCell))

            # Controllo nello sheet RNS DAtaset-1 che le colonne bce non siano vuote
            worksheet=doc["RNC Dataset-1"]
            if worksheet["B3"].value==None:
                self.errInCDR=True
                self.listLineerrInCDR.append("The cell in \"rncId\"(B3) in sheet \"RNC Dataset-1\" can not be empty\n\n")
            if worksheet["C3"].value==None:
                self.errInCDR=True
                self.listLineerrInCDR.append("The cell in \"NODE NAME\"(C3) in sheet \"RNC Dataset-1\" can not be empty\n\n")
            if worksheet["E3"].value==None:
                self.errInCDR=True
                self.listLineerrInCDR.append("The cell in \"rbsId\"(E3) in sheet \"RNC Dataset-1\" can not be empty\n\n")

            # sheet RN RNC-RBS Dataset-1:
            # i valori in colonna r devono finire con una lettera tra U-V-Q-R-W-P e un numero che va da 1 a nSector
            # i valori della colonna q non devono ripetersi
            # i valori della colonna s non devono ripetersi
            # i valore nella colonna r non deve essere presente nella colonna an ma devono essere presenti tutte le altre celle dello stesso settore solo se sono state create ? da chiedere se solo per w va fatto il controllo
            # il valore nella colonna ag deve essere uguale alla cella dello stesso settore che è nello sheet external gsm. se lo sheet externale è vuoto anche lo sheet sarà vuoto
            worksheet=doc["RN RNC-RBS Dataset-1"]
            nTotRow=countCol(doc,"RN RNC-RBS Dataset-1", "R", 12 )
            elemInColR=elemInCol(doc, "RN RNC-RBS Dataset-1", "R", nTotRow, 12 )
            elemInColQ=elemInCol(doc, "RN RNC-RBS Dataset-1", "Q", nTotRow, 12 )
            elemInColS=elemInCol(doc, "RN RNC-RBS Dataset-1", "S", nTotRow, 12 )
            cells=dict(cellP=0, cellQ=0, cellR=0, cellU=0, cellV=0, cellW=0)

            for row in elemInColR:
                if row[-2]=="P":
                    if int(row[-1])>=1 and int(row[-1])<=int(nSector):
                        cells["cellP"]+=1
                    else:
                        self.errInCDR=True
                        self.listLineerrInCDR.append(f"The last character of \"P cells\", in the column \"CELL\" in sheet \"RN RNC-RBS Dataset-1\", must be between 1 and {nSector}.\n\n")
                elif row[-2]=="Q":
                    if int(row[-1])>=1 and int(row[-1])<=int(nSector):
                        cells["cellQ"]+=1
                    else:
                        self.errInCDR=True
                        self.listLineerrInCDR.append(f"The last character of \"Q cells\", in the column \"CELL\" in sheet \"RN RNC-RBS Dataset-1\", must be between 1 and {nSector}.\n\n")
                elif row[-2]=="R":
                    if int(row[-1])>=1 and int(row[-1])<=int(nSector):
                        cells["cellR"]+=1
                    else:
                        self.errInCDR=True
                        self.listLineerrInCDR.append(f"The last character of \"R cells\", in the column \"CELL\" in sheet \"RN RNC-RBS Dataset-1\", must be between 1 and {nSector}.\n\n")
                elif row[-2]=="U":
                    if int(row[-1])>=1 and int(row[-1])<=int(nSector):
                        cells["cellU"]+=1
                    else:
                        self.errInCDR=True
                        self.listLineerrInCDR.append(f"The last character of \"U cells\", in the column \"CELL\" in sheet \"RN RNC-RBS Dataset-1\", must be between 1 and {nSector}.\n\n")
                elif row[-2]=="V":
                    if int(row[-1])>=1 and int(row[-1])<=int(nSector):
                        cells["cellV"]+=1
                    else:
                        self.errInCDR=True
                        self.listLineerrInCDR.append(f"The last character of \"V cells\", in the column \"CELL\" in sheet \"RN RNC-RBS Dataset-1\", must be between 1 and {nSector}.\n\n")
                elif row[-2]=="W":
                    if int(row[-1])>=1 and int(row[-1])<=int(nSector):
                        cells["cellW"]+=1
                    else:
                        self.errInCDR=True
                        self.listLineerrInCDR.append(f"The last character of \"W cells\", in the column \"CELL\" in sheet \"RN RNC-RBS Dataset-1\", must be between 1 and {nSector}.\n\n")
                else:
                    self.errInCDR=True
                    self.listLineerrInCDR.append(f"In the column \"CELL\" in sheet \"RN RNC-RBS Dataset-1\" can exist only cells with U-V-Q-R-W-P. There is a wrong cell, {row}.")

            if cells["cellP"]!=0 and cells["cellP"]!=int(nSector):
                self.errInCDR=True
                self.listLineerrInCDR.append(f"The \"P\" cells, in the column \"Cell\" in sheet \"RN RNC-RBS Dataset-1\",  are smaller than the number of sectors which is {nSector}\n\n")
            if cells["cellQ"]!=0 and cells["cellQ"]!=int(nSector):
                self.errInCDR=True
                self.listLineerrInCDR.append(f"The \"Q\" cells, in the column \"Cell\" in sheet \"RN RNC-RBS Dataset-1\",  are smaller than the number of sectors which is {nSector}\n\n")
            if cells["cellR"]!=0 and cells["cellR"]!=int(nSector):
                self.errInCDR=True
                self.listLineerrInCDR.append(f"The \"R\" cells, in the column \"Cell\" in sheet \"RN RNC-RBS Dataset-1\",  are smaller than the number of sectors which is {nSector}\n\n")
            if cells["cellU"]!=0 and cells["cellU"]!=int(nSector):
                self.errInCDR=True
                self.listLineerrInCDR.append(f"The \"U\" cells, in the column \"Cell\" in sheet \"RN RNC-RBS Dataset-1\",  are smaller than the number of sectors which is {nSector}\n\n")
            if cells["cellV"]!=0 and cells["cellV"]!=int(nSector):
                self.errInCDR=True
                self.listLineerrInCDR.append(f"The \"V\" cells, in the column \"Cell\" in sheet \"RN RNC-RBS Dataset-1\",  are smaller than the number of sectors which is {nSector}\n\n")
            if cells["cellW"]!=0 and cells["cellW"]!=int(nSector):
                self.errInCDR=True
                self.listLineerrInCDR.append(f"The \"W\" cells, in the column \"Cell\" in sheet \"RN RNC-RBS Dataset-1\",  are smaller than the number of sectors which is {nSector}\n\n")

            occurenceInQ=diffValuesInSameCol(elemInColQ)
            if len(occurenceInQ)>0:
                self.errInCDR=True
                for occ in occurenceInQ:
                    self.listLineerrInCDR.append(f"The value \"{occ}\" in coloumn \"localCellId\" in sheet \"RN RNC-RBS Dataset-1\", are repeated more than once.\n\n")

            occurenceInS=diffValuesInSameCol(elemInColS)
            if len(occurenceInS)>0:
                self.errInCDR=True
                for occ in occurenceInS:
                    self.listLineerrInCDR.append(f"The value \"{occ}\" in coloumn \"localCellId\" in sheet \"RN RNC-RBS Dataset-1\", are repeated more than once.\n\n")

            for row in range(12, nTotRow+12):
                valColR=worksheet[f"R{row}"].value
                valColAN=worksheet[f"AN{row}"].value
                splitValColAN=valColAN.split(",")
                # print(splitValColAN)

                if valColR in splitValColAN:
                    self.errInCDR=True
                    self.listLineerrInCDR.append(f"The cell \"{valColR}\"(R{row}), in sheet \"RN RNC-RBS Dataset-1\", is present in \"utranCellRef\"(AN{row}).\n\n")

                sectorNumber=valColR[-1]
                for cell in splitValColAN:
                    if cell!=" ":
                        # print(cell)
                        # print(cell[-1], sectorNumber)
                        if cell[-1]!=sectorNumber:
                            self.errInCDR=True
                            self.listLineerrInCDR.append(f"There are cells of the different sectors in \"utranCellRef\"(AN{row}).\n\n")
                        if not(cell[-2] in lettercellCreated):
                            self.errInCDR=True
                            self.listLineerrInCDR.append(f"In the \"utranCellRef\"(AN{row}) the cell {cell} is not created in column \"CELL\"(R), in sheet \"RN RNC-RBS Dataset-1\".\n\n")

            worksheetExtGsm=doc["External GSM Dataset-1"]
            if worksheetExtGsm[f"C13"].value==None:
                for row in range(12, nTotRow+12):
                    if worksheet[f"AG{row}"].value!="NULL":
                        cell=worksheet[f"AG{row}"].value
                        nameCol=worksheet[f"AG{11}"].value
                        self.errInCDR=True
                        self.listLineerrInCDR.append(f"The value {cell}, in {nameCol} in sheet \"RN RNC-RBS Dataset-1\", is wrong. Should be NULL.\n\n")
            else:
                nTotRowExtGsm=countCol(doc,"External GSM Dataset-1", "F", 13 )
                elemInColFExtGsm=[None]*3
                for row in range(13, nTotRowExtGsm + 13):
                    if not(worksheetExtGsm[f"F{row}"].value in elemInColFExtGsm):
                        if int(worksheetExtGsm[f"F{row}"].value[-1])==1:
                            if elemInColFExtGsm[0]==None:
                                elemInColFExtGsm[0]=worksheetExtGsm[f"F{row}"].value
                            else:
                                if elemInColFExtGsm[0][-2]=="D"and worksheetExtGsm[f"F{row}"].value[-2]=="G":
                                    elemInColFExtGsm[0]=worksheetExtGsm[f"F{row}"].value
                                    pass
                        elif int(worksheetExtGsm[f"F{row}"].value[-1])==2:
                            if elemInColFExtGsm[1]==None:
                                elemInColFExtGsm[1]=worksheetExtGsm[f"F{row}"].value
                            else:
                                if elemInColFExtGsm[1][-2]=="D"and worksheetExtGsm[f"F{row}"].value[-2]=="G":
                                    elemInColFExtGsm[1]=worksheetExtGsm[f"F{row}"].value
                                    pass
                        elif int(worksheetExtGsm[f"F{row}"].value[-1])==3:
                            if elemInColFExtGsm[2]==None:
                                elemInColFExtGsm[2]=worksheetExtGsm[f"F{row}"].value
                            else:
                                if elemInColFExtGsm[2][-2]=="D"and worksheetExtGsm[f"F{row}"].value[-2]=="G":
                                    elemInColFExtGsm[2]=worksheetExtGsm[f"F{row}"].value
                                    pass
                        # elemInColFExtGsm.append(worksheetExtGsm[f"F{row}"].value)

                for row in range(12, nTotRow+12):
                    if int(worksheet[f"R{row}"].value[-1])==1:
                        if worksheet[f"AG{row}"].value!=elemInColFExtGsm[0]:
                            cell=worksheet[f"AG{row}"].value
                            nameCol=worksheet[f"AG{11}"].value
                            self.errInCDR=True
                            self.listLineerrInCDR.append(f"The value {cell}, in {nameCol} in sheet \"RN RNC-RBS Dataset-1\", is wrong. Should be {elemInColFExtGsm[0]}\n\n")
                            # print(f"The value {cell}, in {nameCol} in sheet \"RN RNC-RBS Dataset-1\", is wrong. Should be {elemInColFExtGsm[0]}")
                    if int(worksheet[f"R{row}"].value[-1])==2:
                        if worksheet[f"AG{row}"].value!=elemInColFExtGsm[1]:
                            cell=worksheet[f"AG{row}"].value
                            nameCol=worksheet[f"AG{11}"].value
                            self.errInCDR=True
                            self.listLineerrInCDR.append(f"The value {cell}, in {nameCol} in sheet \"RN RNC-RBS Dataset-1\", is wrong. Should be {elemInColFExtGsm[1]}\n\n")
                            # print(f"The value {cell}, in {nameCol} in sheet \"RN RNC-RBS Dataset-1\", is wrong. Should be {elemInColFExtGsm[1]}")
                    if int(worksheet[f"R{row}"].value[-1])==3:
                        if worksheet[f"AG{row}"].value!=elemInColFExtGsm[2]:
                            cell=worksheet[f"AG{row}"].value
                            nameCol=worksheet[f"AG{11}"].value
                            self.errInCDR=True
                            self.listLineerrInCDR.append(f"The value {cell}, in {nameCol} in sheet \"RN RNC-RBS Dataset-1\", is wrong. Should be {elemInColFExtGsm[2]}\n\n")
                            # print(f"The value {cell}, in {nameCol} in sheet \"RN RNC-RBS Dataset-1\", is wrong. Should be {elemInColFExtGsm[2]}")

            #sheet EutranFreqRelation-1
            # nella colonna b devono essere presenti solo le celle create in Cell(R) in rn rnc dataset-1
            worksheet=doc["EutranFreqRelation-1"]
            nTotRow=countCol(doc,"EutranFreqRelation-1", "B", 2 )
            for row in range(2, nTotRow+2):
                if not(worksheet[f"B{row}"].value in cellCreated):
                    cell=worksheet[f"B{row}"].value
                    self.errInCDR=True
                    self.listLineerrInCDR.append(f"The cell {cell}, in \"UTRANCELL\"(B{row}) in sheet \"EutranFreqRelation-1\", is not created in column \"CELL\"(R), in sheet \"RN RNC-RBS Dataset-1\".\n\n")

            # sheet RN RNC neighbour U2U Dataset-1:
            # le colonne B e E devono avere la stessa stringa che deve essere uguale a alla colonna C del RNC Dataset-1 ;
            # nelle colonne C e D devono comparire SOLO le celle dichiarate in RN RNC-RBS Dataset-1 colonna R.
            worksheet=doc["RN RNC neighbour U2U Dataset-1"]
            nTotRow=countCol(doc,"RN RNC neighbour U2U Dataset-1", "B", 2 )

            cellCreatedInRNcRBS=[]
            nTotRowInRNcRBS=countCol(doc,"RN RNC-RBS Dataset-1", "R", 12 )
            cellCreatedInRNcRBS=elemInCol(doc, "RN RNC-RBS Dataset-1", "R", nTotRowInRNcRBS, 12)

            for row in range(2, nTotRow+2):
                if worksheet[f"B{row}"].value!=rncDataset:
                    cell=worksheet[f"B{row}"].value
                    nameCol=worksheet[f"B{1}"].value
                    self.errInCDR=True
                    self.listLineerrInCDR.append(f"The value {cell}, in {nameCol} in sheet \"RN RNC neighbour U2U Dataset-1\", can not be different from the value in \"NODE NAME\" in \"RNC Dataset-1\".\n\n")
                if worksheet[f"E{row}"].value!=rncDataset:
                    cell=worksheet[f"E{row}"].value
                    nameCol=worksheet[f"E{1}"].value
                    self.errInCDR=True
                    self.listLineerrInCDR.append(f"The value {cell}, in {nameCol} in sheet \"RN RNC neighbour U2U Dataset-1\", can not be different from the value in \"NODE NAME\" in \"RNC Dataset-1\".\n\n")
                if not(worksheet[f"C{row}"].value in cellCreatedInRNcRBS):
                    cell=worksheet[f"C{row}"].value
                    nameCol=worksheet[f"C{1}"].value
                    self.errInCDR=True
                    self.listLineerrInCDR.append(f"The cell {cell}, in {nameCol} in sheet \"RN RNC neighbour U2U Dataset-1\", can not exists in \"CELL\" in \"RN RNC-RBS Dataset-1\".\n\n")
                if not(worksheet[f"D{row}"].value in cellCreatedInRNcRBS):
                    cell=worksheet[f"D{row}"].value
                    nameCol=worksheet[f"D{1}"].value
                    self.errInCDR=True
                    self.listLineerrInCDR.append(f"The cell {cell}, in {nameCol} in sheet \"RN RNC neighbour U2U Dataset-1\", can not exists in \"CELL\" in \"RN RNC-RBS Dataset-1\".\n\n")

            # sheet RN RNC neighbour U2G Dataset-1:
            # la colonna B deve essere uguale alla colonna C del RNC Dataset-1:
            # nella colonna C ci devono essere SOLO (anche ripetute) le celle definite in colonna R RN RNC-RBS Dataset-1;
            # nella colonna D ci devono essere SOLO (anche ripetute) le celle definite in colonna F dello sheet External GSM Dataset-1. da finire
            worksheet=doc["RN RNC neighbour U2G Dataset-1"]
            nTotRow=countCol(doc,"RN RNC neighbour U2G Dataset-1", "C", 2 )

            cellCreatedInExtGsm=[]
            nTotRowInExtGsm=countCol(doc,"External GSM Dataset-1", "F", 12 )
            cellCreatedInExtGsm=elemInCol(doc, "RN RNC-RBS Dataset-1", "F", nTotRowInRNcRBS, 12)

            for row in range(2, nTotRow+2):
                if worksheet[f"B{row}"].value!=rncDataset:
                    cell=worksheet[f"B{row}"].value
                    nameCol=worksheet[f"B{11}"].value
                    self.errInCDR=True
                    self.listLineerrInCDR.append(f"The value {cell}, in {nameCol} in sheet \"RN RNC neighbour U2G Dataset-1\", can not be different from the value in \"NODE NAME\" in \"RNC Dataset-1\".\n\n")
                if not(worksheet[f"C{row}"].value in cellCreatedInRNcRBS):
                    cell=worksheet[f"C{row}"].value
                    nameCol=worksheet[f"C{1}"].value
                    self.errInCDR=True
                    self.listLineerrInCDR.append(f"The cell {cell}, in {nameCol} in sheet \"RN RNC neighbour U2G Dataset-1\", can not exists in \"CELL\" in \"RN RNC-RBS Dataset-1\".\n\n")


            # sheet External GSM Dataset-1:
            # se la colonna F è non-vuota, allora la colonna C deve essere uguale alla colonna C del RNC Dataset-1;
            # se la colonna F è non-vuota, allora nella colonna G i valori NON si devono ripetere;
            # se la colonna F è non-vuota, allora la colonna H deve essere sempre 7
            # se la colonna F è non-vuota, allora I può assumere valori da 0 a 7
            worksheet=doc["External GSM Dataset-1"]
            nTotRow=countCol(doc,"External GSM Dataset-1", "C", 13 )

            if nTotRow != 0:
                for row in range(13, nTotRow+13):
                    if worksheet[f"F{row}"].value != None:
                        if worksheet[f"C{row}"].value!=rncDataset:
                            cell=worksheet[f"C{row}"].value
                            nameCol=worksheet[f"C{11}"].value
                            self.errInCDR=True
                            self.listLineerrInCDR.append(f"The value {cell}, in {nameCol} in sheet \"External GSM Dataset-1\", can not be different from the value in \"NODE NAME\" in \"RNC Dataset-1\".\n\n")
                            if int(worksheet[f"H{row}"].value)!=7:
                                cell=worksheet[f"H{row}"].value
                                nameCol=worksheet[f"H{11}"].value
                                self.errInCDR=True
                                self.listLineerrInCDR.append(f"The value {cell}, in {nameCol} in sheet \"External GSM Dataset-1\", must be 7.\n\n")
                                if int(worksheet[f"I{row}"].value)<0 or int(worksheet[f"I{row}"].value)>7:
                                    cell=worksheet[f"I{row}"].value
                                    nameCol=worksheet[f"I{11}"].value
                                    self.errInCDR=True
                                    self.listLineerrInCDR.append(f"The value {cell}, in {nameCol} in sheet \"External GSM Dataset-1\", must be between 0 and 7.\n\n")

                occurenceInG=occurenceInCol(doc, "External GSM Dataset-1", "G", 13)
                if len(occurenceInG)>0:
                    for element in occurenceInG:
                        nameCol=worksheet[f"G{11}"].value
                        self.errInCDR=True
                        self.listLineerrInCDR.append(f"In the column {nameCol} in sheet \"External GSM Dataset-1\", the value {element[0]} is present more than once.\n\n")

        except Exception as e:
            self.errInCDR=None
            self.exc_type, self.exc_obj, self.exc_tb = sys.exc_info()
            self.fileWithGenericError=path.split("/")[-1]
            messagebox.showerror(message=f"Generic error in Tool with {self.fileWithGenericError}.\nError type: {self.exc_type}\nError line: {self.exc_tb.tb_lineno}")

    def lte(self, path):
        try:
            doc= openpyxl.load_workbook(path, data_only=True)
            cellCreated=[]

            worksheet=doc["EutranCell"]
            nTotRow=countCol(doc,"EutranCell", "B", 12 )
            for row in range(12, nTotRow+12):
                cellCreated.append(worksheet[f"B{row}"].value)

            # sheet EutranCell:
            # colonna A le celle devono finire con una lettera tra a, b, c, f e un numero
            # LA colonna i non deve avere valori ripetuti
            # i valoei di j possono avere valore compreso tra 0e 167
            # i valori di k possono variare tra 0 e 2
            # Se nella colonna j ci sono valori uguali, nelle rispettive celle della colonna k non devono esserci valori uguali
            worksheet=doc["EutranCell"]
            nTotRow=countCol(doc,"EutranCell", "B", 12 )
            letterAllowedForCell=["A", "B", "C", "F"]
            for row in range(12, nTotRow+12):
                if worksheet[f"B{row}"].value[-2] in letterAllowedForCell:
                    if not(int(worksheet[f"B{row}"].value[-1])):
                        cell=worksheet[f"B{row}"].value
                        nameCol=worksheet[f"B11"].value
                        self.errInCDR=True
                        self.listLineerrInCDR.append(f"The last character of cell {cell}, in \"{nameCol}\" in sheet \"EutranCell\", must be a number\n\n")
                else:
                    cell=worksheet[f"B{row}"].value
                    nameCol=worksheet[f"B11"].value
                    self.errInCDR=True
                    self.listLineerrInCDR.append(f"The cell {cell}, in \"{nameCol}\"(B{row}) in sheet \"EutranCell\", must be only \"A\", \"B\", \"C\", \"F\"\n\n")

                if int(worksheet[f"J{row}"].value)<0 or int(worksheet[f"J{row}"].value)>167:
                    cell=worksheet[f"J{row}"].value
                    nameCol=worksheet[f"J{11}"].value
                    self.errInCDR=True
                    self.listLineerrInCDR.append(f"The value {cell}, in \"{nameCol}\"(J{row}) in sheet \"EutranCell\", must be between 0 and 167 \n\n")

                if int(worksheet[f"K{row}"].value)<0 or int(worksheet[f"K{row}"].value)>2:
                    cell=worksheet[f"K{row}"].value
                    nameCol=worksheet[f"J{11}"].value
                    self.errInCDR=True
                    self.listLineerrInCDR.append(f"The value {cell}, in \"{nameCol}\"(K{row}) in sheet \"EutranCell\", must be between 0 and 2 \n\n")

            elemInColI=elemInCol(doc, "EutranCell", "I", nTotRow, 12)
            occurenceInI=diffValuesInSameCol(elemInColI)
            if len(occurenceInI)>0:
                self.errInCDR=True
                for occ in occurenceInI:
                    self.listLineerrInCDR.append(f"The value \"{occ}\" in coloumn \"CellId\" in sheet \"EutranCell\", are repeated more than once.\n\n")

            # elemInColJ=elemInCol(doc, "EutranCell", "J", nTotRow, 12)
            # occurence=[]
            for row in range(12, nTotRow+12):
                firstOcc=worksheet[f"J{row}"].value
                rowFirstOcc=row
                valColKFirsOcc=worksheet[f"K{row}"].value
                valColPFirsOcc=worksheet[f"P{row}"].value

                for row in range(12, nTotRow+12):
                    if worksheet[f"J{row}"].value==firstOcc and row!=rowFirstOcc:
                        valcolKrow=worksheet[f"K{row}"].value
                        valcolProw=worksheet[f"P{row}"].value
                        if valcolProw==valColPFirsOcc:
                            if valColKFirsOcc==valcolKrow:
                                # cell=worksheet[f"{row}"].value
                                nameCol=worksheet[f"K{11}"].value
                                self.errInCDR=True
                                self.listLineerrInCDR.append(f"The value \"{valColKFirsOcc}\"(K{rowFirstOcc}) and \"{valcolKrow}\"(K{row}), in \"{nameCol}\" in sheet \"EutranCell, can not be the same.\"\n\n")

            # sheet EUtranFreqRelation:
            # nella colonna b devono esserci solo i valori di cellCreated
            worksheet=doc["EUtranFreqRelation"]
            nTotRow=countCol(doc,"EUtranFreqRelation", "B", 12 )
            for row in range(12, nTotRow+12):
                if not(worksheet[f"B{row}"].value in cellCreated):
                    cell=worksheet[f"B{row}"].value
                    nameCol=worksheet[f"B{11}"].value
                    self.errInCDR=True
                    self.listLineerrInCDR.append(f"The cell \"{cell}\", in \"{nameCol}\"(B{row}) in sheet \"EUtranFreqRelation\", is not created in \"Cell Name\" in sheet \"EutranCell\".\n\n")

            # sheet UtranFreqRelation:
            # nella colonna b devono esserci solo i valori di cellCreated
            worksheet=doc["UtranFreqRelation"]
            nTotRow=countCol(doc,"UtranFreqRelation", "B", 12 )
            for row in range(12, nTotRow+12):
                if not(worksheet[f"B{row}"].value in cellCreated):
                    cell=worksheet[f"B{row}"].value
                    nameCol=worksheet[f"B{11}"].value
                    self.errInCDR=True
                    self.listLineerrInCDR.append(f"The cell \"{cell}\", in \"{nameCol}\"(B{row}) in sheet \"UtranFreqRelation\", is not created in \"Cell Name\" in sheet \"EutranCell\".\n\n")

            # sheet EUtranCellRelation:
            # nella colonna b devono esserci solo i valori di cellCreated
            # nella colonna c devono esserci solo i valori di cellCreated
            # lo split della colonna e deve contenere solo i valori di cellCreated
            worksheet=doc["EUtranCellRelation"]
            nTotRow=countCol(doc,"EUtranCellRelation", "B", 12 )
            for row in range(12, nTotRow+12):
                if not(worksheet[f"B{row}"].value in cellCreated):
                    cell=worksheet[f"B{row}"].value
                    nameCol=worksheet[f"B{11}"].value
                    self.errInCDR=True
                    self.listLineerrInCDR.append(f"The cell \"{cell}\", in \"{nameCol}\"(B{row}) in sheet \"EUtranCellRelation\", is not created in \"Cell Name\" in sheet \"EutranCell\".\n\n")

                if not(worksheet[f"C{row}"].value in cellCreated):
                    cell=worksheet[f"C{row}"].value
                    nameCol=worksheet[f"C{11}"].value
                    self.errInCDR=True
                    self.listLineerrInCDR.append(f"The cell \"{cell}\", in \"{nameCol}\"(B{row}) in sheet \"EUtranCellRelation\", is not created in \"Cell Name\" in sheet \"EutranCell\".\n\n")

                splitValColE=worksheet[f"E{row}"].value.split("-")
                for string in splitValColE:
                    if not(string in cellCreated):
                        cell=worksheet[f"E{row}"].value
                        nameCol=worksheet[f"E{11}"].value
                        self.errInCDR=True
                        self.listLineerrInCDR.append(f"In the {nameCol}(E{row}) in sheet \"EutranCell\", can exist only cells created in \"Cell Name\" in sheet \"EutranCell\".\n\n")
                        pass
        except Exception as e:
            self.errInCDR=None
            self.exc_type, self.exc_obj, self.exc_tb = sys.exc_info()
            self.fileWithGenericError=path.split("/")[-1]
            messagebox.showerror(message=f"Generic error in Tool with {self.fileWithGenericError}.\nError type: {self.exc_type}\nError line: {self.exc_tb.tb_lineno}")

def countCol(excel,sheet, col,startRow):

    i=startRow
    nTotCol=0
    while excel[sheet][col+str(i)].value!=None:
        nTotCol+=1
        i+=1
        pass

    return nTotCol
    pass

def countRowFill(excel, row, startCol):

    startColList=list(startCol)
    colFill=0

    while "".join(startColList)!="AL":
        if excel["BTS"]["".join(startColList)+str(row)].value!=None:
            colFill+=1
        startColList[-1]=chr(ord(startColList[-1])+1)
    return colFill
    pass
# ritorna una lisa di tutti gli elementi della colonna, prenedno il documento, losheet, la colonna, il numero totale di colonne e la colonna da cui iniziare
def elemInCol(excel,sheet, col, nTotCol, rowStart):
    elementCol=[]
    for element in range(rowStart,nTotCol+rowStart):
        elementCol.append(excel[sheet][col+str(element)].value)
        pass
    return elementCol
    pass
# cerca nella colonna se i valori sono diversi. Se non sono diversi torna errore
def equalValuesInSameCol(doc, sheet, col, rowStart):
    err=[]

    for column in col:
        cont=rowStart
        diff=False
        valToCheck=doc[sheet][str(column)+str(cont)].value
        while doc[sheet][str(column)+str(cont)].value!=None:
            if doc[sheet][str(column)+str(cont)].value!= valToCheck:
                diff=True
            cont+=1
        if diff:
            err.append(str(column))
    return err

    pass

def occurenceInCol(doc, sheet, col, rowStart):
    nTotRow=countCol(doc, sheet, col, rowStart)
    elemInColX=elemInCol(doc, sheet, col, nTotRow, rowStart)
    occurence=[]
    for e in elemInColX:
        occE=elemInColX.count(e)
        if occE>1:
            # occurence.append([e,occE])
            if len(occurence)>0:
                exist=False
                for ele in occurence:
                    if e==ele[0]:
                        exist=True
                        break
                if not(exist):
                    occurence.append([e,occE])

            else:
                occurence.append([e,occE])
    return occurence

def equalValuesInOtherCol(doc, sheet, col, rowStart, sheetToCheck, colToCheck, rowToCheck):
    err=[]
    valToCheck=doc[sheetToCheck][str(colToCheck)+str(rowToCheck)].value
    cont=rowStart
    diff=False

    while doc[sheet][str(col)+str(cont)].value!=None:
        if doc[sheet][str(col)+str(cont)].value!= valToCheck:
            diff=True
        cont+=1
    if diff:
        return  True
    else:
        return  False
    pass

def removeSpace(uniqueCode):
        tempUniqueCode=uniqueCode.strip(" ")
        nSpace=tempUniqueCode.count(" ")
        if nSpace!=0:
            return tempUniqueCode.replace(" ","", nSpace-1)
        else:
            return tempUniqueCode

def nSectors(uniqueCode):
    tempUniqueCode=uniqueCode.split("_")[1]
    # print(tempUniqueCode)
    if "S" in tempUniqueCode:
        return tempUniqueCode[0]
    else:
        tempUniqueCode=uniqueCode.split("_")[2]
        if "S" in tempUniqueCode:
            return tempUniqueCode[0]
# cerca nella colonna se i valori sono uguali. Se sonon uguali torna errore e quali valori si ripetono
def diffValuesInSameCol(colToAnalize):
    repeatedValues=[]
    occurence=False
    for row in colToAnalize:
        rowToCompare=row
        if colToAnalize.count(row)>1:
            if not(row in repeatedValues):
                repeatedValues.append(row)
                occurence=True
    if occurence:
        return repeatedValues
    else:
        return repeatedValues
